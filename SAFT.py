import sys
import os.path
import platform
import copy
import itertools

#PySide2 imports
from PySide2 import QtCore, QtGui
from PySide2.QtCore import Slot
from PySide2 import __version__ as pyside_version
from PySide2.QtWidgets import QApplication, QMainWindow, QWidget, QGridLayout, QMessageBox, QFileDialog, QAction, QGroupBox, QHBoxLayout, QRadioButton, QDialog, QVBoxLayout, QCheckBox, QButtonGroup, QFrame

#package imports
import numpy as np
import pandas as pd
from scipy import __version__ as scipy_version
import scipy.signal as scsig
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#SAFT imports
from clicker import clickAlgebra
from extractPeakResponses import extractPeaksDialog
from fitHistograms import histogramFitDialog
from processGroupedPeaks import groupPeakDialog
from quantal import fit_nGaussians, nGaussians_display
from baselines import savitzky_golay, baseline_als, baselineIterator
from dataStructures import Store, Dataset, Results, HistogramsR
from helpMessages import gettingStarted
import utils            #addFileSuffix, findCurve, findScatter etc


#Import pg last to avoid namespace-overwrite problems?
import pyqtgraph as pg


class QHLine(QFrame):
    ### from https://stackoverflow.com/questions/5671354
    def __init__(self):
        super(QHLine, self).__init__()
        self.setFrameShape(QFrame.HLine)
        self.setFrameShadow(QFrame.Sunken)

class SAFTMainWindow(QMainWindow):
    
    ### Methods
    ###
    ### createMenu                  : make menubar
    ### about                       : About the app
    ### getStarted                  : a help file for novices
    ### createSplitTraceLayout      : for display of split traces
    ### createPlotWidgets           : build the plots
    ### createLinearRegion          : linear selection for zoom window
    ### mouseMoved                  : when the mouse moves in zoom
    ### splitState                  : change the split state for the overview window
    ### manualPeakToggle            :
    ###
    ###
    ###
    
    def __init__(self, *args, **kwargs):
        super(SAFTMainWindow, self).__init__(*args, **kwargs)
        
        self.setWindowTitle("Semi-Automatic Fluorescence Trace analysis")
        self.central_widget = QWidget()
        self.central_layout = QGridLayout()
        self.central_widget.setLayout(self.central_layout)
        self.setCentralWidget(self.central_widget)
        self.resize(1500,800)           # works well on MacBook Retina display
        
        self.split_traces = False
        self.LR_created = False                 # was a pg linear region created yet?
        self.wasManualOnce = False              # was manual editing of peaks ever engaged?
        self.simplePeaks = False                # choice of peak finding algorithm
        self.autoPeaks = True                   # find peaks automatically or manually
        self.cwt_width = 5                      # width of the continuous wavelet transform peak finding
        
        self.store = Store()                    # store for the datasets not being analysed
        self.dataLoaded = False                 # was any data loaded yet?
        self.pauseUpdates = False               # should updates be paused whilst we make a lot of changes to GUI?
        self.noCrosshair = True                 # is there any crosshair shown?
        self.workingDataset = Dataset("Empty")  # unnamed, empty dataset for traces, pk results and GUI settings
        self.workingDataset.ROI_list = None
        self.filename = None
        
        self.conditions = []                    # conditions will be sheet names from xlsx
        self.datasetList_CBX = ['-']            # maintain our own list of datasets from the combobox
        self.extPa = {}                         # external parameters for the peak scraping dialog
        self.dataLock = True                    # when manual peak editing, lock to trace data
        self.noPeaks = True                     # were any peaks found yet?
        self.fitHistogramsOption = False        # histograms are not fitted by default, checkbox -> False later
        self.saveHistogramsOption = False       # histograms are not saved by default,  checkbox -> False later
        
        # setup main window widgets and menus
        self.createPlotWidgets()
        self.createControlsWidgets()
        self.createMenu()
        self.toggleDataSource = False
        
        
    def createMenu(self):
        # Skeleton menu commands
        self.file_menu = self.menuBar().addMenu("File")
        self.analysis_menu = self.menuBar().addMenu("Analysis")
        self.help_menu = self.menuBar().addMenu("Help")
        
        self.file_menu.addAction("About FSTPA", self.about) #this actually goes straight into the FSTPA menu
        self.file_menu.addAction("Open File", self.open_file)
        self.file_menu.addAction("Save Peaks", self.save_peaks)
        
        self.file_menu.addAction("Save baselined", self.save_baselined)
        
        self.analysis_menu.addAction("Extract all peaks", self.extractAllPeaks)
        self.analysis_menu.addAction("Grouped peak stats", self.getGroups)
        self.analysis_menu.addAction("Quantal Histogram Fit", self.launchHistogramFit)
        
        self.help_menu.addAction("Getting Started", self.getStarted)
    
    
    def about(self):
        QMessageBox.about (self, "About SAFT",
        """ ----*- SAFT {0} -*----
        \nSemi-Automatic Fluorescence Trace analysis
        \nAndrew Plested FMP- and HU-Berlin 2020
        \nThis application can analyse sets of fluorescence time series.
        \nIt makes heavy use of PyQtGraph ({8}, Luke Campagnola).
        \nPython {1}
        \nPandas {2}, Numpy {3}, SciPy {4}
        \nPySide2 {5} built on Qt {6}
        \nRunning on {7}
        """.format(__version__, platform.python_version(), pd.__version__, np.__version__, scipy_version, pyside_version, QtCore.__version__, platform.platform(), pg.__version__))
    
    def modalWarning(self, s):
        #print("click", s)


        QMessageBox.warning(self, 'Warning', s)
        #dlg.setWindowTitle("HELLO!")
        #dlg.exec_()
    
    
    def getStarted(self):
        
        QMessageBox.information(self, "Getting Started", gettingStarted())
    
    
    def createSplitTraceLayout(self):
        """Optional split view with each ROI trace in a separate plot (not the default)"""
        
        # Store the plot items in a list - can't seem to get them easily otherwise?
        data = []
        self.p1stackMembers = []
        for c in self.conditions:
            memberName = c + " trace"
            p1_stack_member = self.p1stack.addPlot(title=c, y=data, name=memberName)
            p1_stack_member.hideAxis('bottom')
            self.p1stackMembers.append(p1_stack_member)
            self.p1stack.nextRow()
            #print (c, len(self.p1stackMembers))
        
        #link y-axes - using final member of stack as anchor
        for s in self.p1stackMembers:
            if s != p1_stack_member:
                s.setXLink(p1_stack_member)
                s.setYLink(p1_stack_member)
                
        #add back bottom axis to the last graph
        p1_stack_member.showAxis("bottom")
        p1_stack_member.setLabel('bottom', "Time (s)")
    
    
    def createPlotWidgets(self):
        """analysis plots"""
        
        # traces plot
        data = []
        self.plots = pg.GraphicsLayoutWidget()
        self.p1rc = (1,0)
        self.p1 = self.plots.addPlot(y=data, row=self.p1rc[0], col=self.p1rc[1], rowspan=3, colspan=1)
        self.p1.setTitle(title="Traces and background subtraction", color="F0F0F0", justify="right")
        self.p1.setLabel('left', "dF / F")
        self.p1.setLabel('bottom', "Time (s)")
        self.p1.vb.setLimits(xMin=0)
        #just a blank for now, populate after loading data to get the right number of split graphs
        self.p1stack = pg.GraphicsLayout()
        
        if self.dataLoaded:
            createLinearRegion()
        
        # Histograms
        self.p2 = self.plots.addPlot(row=0, col=0, rowspan=1, colspan=1)
        self.p2.setTitle("Peak Histograms", color="F0F0F0", justify="right")
        self.p2.setLabel('left', "N")
        self.p2.setLabel('bottom', "dF / F")
        self.p2.vb.setLimits(xMin=0, yMin=0)
        self.p2.addLegend(offset=(-50,50))
        #self.p2.legend.setOffset()
        
        # zoomed editing region , start in auto peak mode
        self.p3 = self.plots.addPlot(y=data, row=0, col=1, rowspan=4, colspan=2)
        self.p3.setTitle('Zoom - auto peak mode', color="F0F0F0", justify="right")
        self.p3.setLabel('left', "dF / F")
        self.p3.setLabel('bottom', "Time (s)")
        self.p3.setFixedWidth(550)
        self.p3vb = self.p3.vb
        
        # draw the crosshair if we are in manual editing mode
        self.p3proxyM = pg.SignalProxy(self.p3.scene().sigMouseMoved, rateLimit=60, slot=self.mouseMoved)
        
        # what does this do??
        self.p3.scene().sigMouseClicked.connect(self.clickRelay)
        #self.p3.sigMouseClicked.connect(self.clickRelay)
        
        self.plots.cursorlabel = pg.LabelItem(text='', justify='right')
        
        # to stop label jiggling about (graphicswidget method)
        self.plots.cursorlabel.setFixedWidth(100)
        
        self.plots.peakslabel = pg.LabelItem(text='', justify='left')
        
        # to stop label jiggling about (graphicswidget method)
        self.plots.peakslabel.setFixedWidth(100)
        
        self.plots.addItem(self.plots.cursorlabel, row=4, col=2)
        self.plots.addItem(self.plots.peakslabel, row=4, col=1)
        
        self.central_layout.addWidget(self.plots, row=0, col=0, rowspan=1,colspan=2)
     
     
    def clickRelay(self, *args):
        """Logic to avoid the click signal getting sent out if manual peak editing is not on."""
        if self.autoPeaks:
            print ("Turn on manual peak editing to get some value for your clicks.\nFor debugging: ", args)
            self.modalWarning ("Turn on manual peak editing to get some value for your clicks.\nFor debugging: {}".format(args))
            return
        else:
            # the asterisk in call unpacks the tuple into individual arguments.
            # a free click should be locked to the data (as the crosshair is)
            self.cA.onClick (*args, dataLocked=self.dataLock)
    
    
    def createLinearRegion(self):
        """Linear region in p1 that defines the x-region in p3 (manual editing window)"""
        # taken from pyqtgraph examples.
        
        if self.LR_created == False:
            self.LR_created = True
            xrange = self.ranges['xmax']-self.ranges['xmin']
            self.lr = pg.LinearRegionItem([xrange/2, xrange/1.5])
            self.lr.setZValue(-10)
    
        def updatePlot():
            self.p3.setXRange(*self.lr.getRegion(), padding=0)
        def updateRegion():
            self.lr.setRegion(self.p3.getViewBox().viewRange()[0])
        
        self.lr.sigRegionChanged.connect(updatePlot)
        self.p3.sigXRangeChanged.connect(updateRegion)
        
        if self.split_traces:
            for s in self.p1stackMembers:
                s.addItem(self.lr)
        else:
            self.p1.addItem(self.lr)
        updatePlot()
    
    
    def mouseMoved(self, evt):
        """Crosshair in p3 shown during manual fitting"""
        if self.autoPeaks == False:
            pos = evt[0]  ## using signal proxy turns original arguments into a tuple
            if self.p3.sceneBoundingRect().contains(pos):
                mousePoint = self.p3vb.mapSceneToView(pos)
                
                # there should be two plot data items, find the curve data
                _c = utils.findCurve(self.p3.items)
                sx, sy = _c.getData()
            
                # quantize x to curve, and get corresponding y that is locked to curve
                idx = np.abs(sx - mousePoint.x()).argmin()
                ch_x = sx[idx]
                ch_y = sy[idx]
                self.hLine.setPos(ch_y)
                self.vLine.setPos(ch_x)
                
                # print ("update label: x={:.2f}, y={:.2f}".format(ch_x, ch_y))
                self.plots.cursorlabel.setText("Cursor: x={: .2f}, y={: .3f}".format(ch_x, ch_y))
    
    
    def splitState(self, b):
        """Called when trace display selection radio buttons are activated """
        if b.text() == "Split traces":
            if b.isChecked() == True:
                self.split_traces = True
            else:
                self.split_traces = False
            
        if b.text() == "Combined traces":
            if b.isChecked() == True:
                self.split_traces = False
            else:
                self.split_traces = True
        
        tobeRemoved = self.plots.getItem(*self.p1rc)
        print ("Removing", tobeRemoved)
        
        #self.p1rc is a tuple containing the position (row-column) of p1
        if self.split_traces:
            self.plots.removeItem(tobeRemoved)
            self.plots.addItem(self.p1stack, *self.p1rc, 3, 1)
        else:
            self.plots.removeItem(tobeRemoved)
            self.plots.addItem(self.p1, *self.p1rc, 3, 1)

        # call general update method
        self.ROI_Change()
    
    def toggleDataLogic (self, b):
        if b.isChecked() == True:
            self.showExtracted = True
            # some hook to take traces and data from extracted etc
            print ("data from extracted")
        else:
            self.showExtracted = False
            # some hook to take traces and peaks data from searched
            print ("raw data from search")
    
    def saveHistogramsLogic (self, b):
        if b.isChecked() == True:
            self.saveHistogramsOption = True
        else:
            self.saveHistogramsOption = False
    
    def fitHistogramsLogic (self, b):
        if b.isChecked() == True:
            self.fitHistogramsOption = True
        else:
            self.fitHistogramsOption = False
    
    
    def manualPeakToggle (self, b=None):
        """Disable controls if we are editing peaks manually"""
        #print ("MPT {}".format(b))
        if self.manual.isChecked() == True:
            # enter manual mode
            print ("Manual peak editing")
            self.autoPeaks = False
            self.wasManualOnce = True
            
            # disable all controls that could trigger auto peak finding
            self.peak_CB.setDisabled(True)
            self.SGsmoothing_CB.setDisabled(True)
            self.cwt_SNR_Spin.setDisabled(True)
            self.cwt_w_Spin.setDisabled(True)
            self.auto_bs_lam_slider.setDisabled(True)
            self.auto_bs_P_slider.setDisabled(True)
            self.autobs_Box.setDisabled(True)
            self.removeSml_Spin.setDisabled(True)
            
            # Turn on crosshair and change mouse mode in p3.
            if self.noCrosshair:
                self.vLine = pg.InfiniteLine(angle=90, movable=False)
                self.hLine = pg.InfiniteLine(angle=0, movable=False)
                self.p3.addItem(self.vLine, ignoreBounds=True)
                self.p3.addItem(self.hLine, ignoreBounds=True)
                self.noCrosshair = False
            
            # add a hint
            self.p3.setTitle('Zoom - Manual editing  L-click to add/remove peaks', color="F0F0F0", justify="right")
            
        elif self.wasManualOnce:
            # Enter auto peak mode
            print ("Auto peak finding")
            self.autoPeaks = True
            
            # Re-enable all the controls for auto peak finding
            self.peak_CB.setEnabled(True)
            self.SGsmoothing_CB.setEnabled(True)
            self.cwt_SNR_Spin.setEnabled(True)
            self.cwt_w_Spin.setEnabled(True)
            self.auto_bs_lam_slider.setEnabled(True)
            self.auto_bs_P_slider.setEnabled(True)
            self.autobs_Box.setEnabled(True)
            self.removeSml_Spin.setEnabled(True)
            
            # Change the hint
            self.p3.setTitle('Zoom - Auto peak mode', color="F0F0F0", justify="right")
            
            if self.noCrosshair == False:
            # Remove crosshair from p3.
                self.p3.removeItem(self.vLine)
                self.p3.removeItem(self.hLine)
                self.noCrosshair = True
        
        else:
            print ("DEBUG: Fell through manualPeakToggle without updating")
     
    def createControlsWidgets(self):
        """control panel"""
        
        controls = pg.LayoutWidget()
        
        histograms = QGroupBox("Histogram options")
        histGrid = QGridLayout()
        
        NBin_label = QtGui.QLabel("No. of bins")
        NBin_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.histo_NBin_Spin = pg.SpinBox(value=100, step=10, bounds=[0, 250], delay=0)
        self.histo_NBin_Spin.setFixedSize(60, 25)
        self.histo_NBin_Spin.valueChanged.connect(self.updateHistograms)
        
        histMax_label = QtGui.QLabel("dF/F max")
        histMax_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.histo_Max_Spin = pg.SpinBox(value=1, step=0.1, bounds=[0.1, 10], delay=0, int=False)
        self.histo_Max_Spin.setFixedSize(60, 25)
        self.histo_Max_Spin.valueChanged.connect(self.updateHistograms)
        
        #toggle show ROI histogram sum
        histsum_label = QtGui.QLabel("Show histograms")
        histsum_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.sum_hist = pg.ComboBox()
        self.sum_hist.setFixedSize(95,25)
        self.sum_hist.addItems(['Separated','Summed'])
        
        self.sum_hist.currentIndexChanged.connect(self.updateHistograms)
        
        #toggle fitting
        self.fitHistogramsToggle = QCheckBox("Fit Histograms", self)
        self.fitHistogramsToggle.toggled.connect(lambda:self.fitHistogramsLogic(self.fitHistogramsToggle))
        self.fitHistogramsToggle.setChecked(self.fitHistogramsOption)
        
        
        
        #fit parameters
        histnG_label = QtGui.QLabel("No. of Gaussians")
        histnG_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        
        self.histo_nG_Spin = pg.SpinBox(value=5, step=1, bounds=[1,10], delay=0, int=True)
        self.histo_nG_Spin.setFixedSize(60, 25)
        self.histo_nG_Spin.valueChanged.connect(self.updateHistograms)
        
        histq_label = QtGui.QLabel("dF ('q') guess")
        histq_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.histo_q_Spin = pg.SpinBox(value=.05, step=0.01, bounds=[0.01,1], delay=0, int=False)
        self.histo_q_Spin.setFixedSize(60, 25)
        self.histo_q_Spin.valueChanged.connect(self.updateHistograms)
        
        self.saveHistogramsToggle = QCheckBox("Save Histograms", self)
        self.saveHistogramsToggle.setChecked(self.saveHistogramsOption)
        self.saveHistogramsToggle.toggled.connect(lambda:self.saveHistogramsLogic(self.saveHistogramsToggle))
        
        
        
        histGrid.addWidget(histsum_label, 0, 0, 1, 2)
        histGrid.addWidget(self.sum_hist, 0, 2, 1, 3)
        histGrid.addWidget(self.fitHistogramsToggle, 1, 5, 1, 2)
        
        histGrid.addWidget(histnG_label, 2, 4, 1, 2)
        histGrid.addWidget(self.histo_nG_Spin, 2, 6)
        
        histGrid.addWidget(histq_label, 3, 4, 1, 2)
        histGrid.addWidget(self.histo_q_Spin, 3, 6)
                
        histGrid.addWidget(histMax_label, 1, 0)
        histGrid.addWidget(self.histo_Max_Spin, 1, 1, 1, 2)
        
        histGrid.addWidget(NBin_label, 2, 0)
        histGrid.addWidget(self.histo_NBin_Spin, 2, 1, 1, 2)
        
        histGrid.addWidget(self.saveHistogramsToggle, 3, 0, 1, 4)
        
        histograms.setLayout(histGrid)
        
        # Data display options panel
        dataPanel = QGroupBox("Data display and processing")
        dataGrid = QGridLayout()
        self.split_B = QRadioButton("Split traces", self)
        self.combine_B = QRadioButton("Combined traces", self)
        self.combine_B.setChecked(True)
        self.split_B.toggled.connect(lambda:self.splitState(self.split_B))
        self.combine_B.toggled.connect(lambda:self.splitState(self.combine_B))
        
        # load dataset button - for Benni
        loadDataBtn = QtGui.QPushButton('Load traces')
        loadDataBtn.clicked.connect(self.open_file)
        
        # select working dataset
        datasetLabel = QtGui.QLabel("Dataset")
        datasetLabel.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        
        self.datasetCBx = pg.ComboBox()
        self.datasetCBx.setFixedWidth(200)
        self.datasetCBx.setItems(self.datasetList_CBX)
        self.datasetCBx.currentIndexChanged.connect(self.datasetChange)
        
        # selection of ROI trace, or mean, variance etc
        ROIBox_label = QtGui.QLabel("Select ROI")
        ROIBox_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        
        self.ROI_selectBox = QtGui.QComboBox()
        self.ROI_selectBox.addItems(['None'])
        self.ROI_selectBox.currentIndexChanged.connect(self.ROI_Change)
        
        d_divider = QHLine()
        d_divider.setFixedWidth(375)
        
        # launch histogram fitting dialog
        # should be inactive until extraction
        self.fitHistDialogBtn = QtGui.QPushButton('Quantal Histogram Fit')
        self.fitHistDialogBtn.clicked.connect(self.launchHistogramFit)
        self.fitHistDialogBtn.setDisabled(True)
        
        # launch peak extraction wizard dialog
        extractPeaksBtn = QtGui.QPushButton('Extract peaks from all ROIs')
        extractPeaksBtn.clicked.connect(self.extractAllPeaks)
        
        # should be inactive until extraction
        self.extractGroupsDialog_Btn = QtGui.QPushButton('Extract grouped responses')
        self.extractGroupsDialog_Btn.clicked.connect(self.getGroups)
        self.extractGroupsDialog_Btn.setDisabled(True)
        
        reference_label = QtGui.QLabel("Reference condition for pattern extraction")
        reference_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        
        self.refSelection = QtGui.QComboBox()
        self.refSelection.setFixedSize(90, 25)       # only width seems to work
        self.refSelection.addItems(['-'])
        self.refSelection.currentIndexChanged.connect(self.ROI_Change)
        
        _buttonList = [self.fitHistDialogBtn, extractPeaksBtn, self.extractGroupsDialog_Btn]
        bsize = (190, 40)
        for b in _buttonList:
            b.setMinimumSize(*bsize)
        
        dataGrid.addWidget(loadDataBtn, 0, 0, 1, 2)
        dataGrid.addWidget(datasetLabel, 0, 2, 1, 2)
        dataGrid.addWidget(self.datasetCBx, 0, 4, 1, 4)
        dataGrid.addWidget(ROIBox_label, 2, 2, 1, 2)
        dataGrid.addWidget(self.ROI_selectBox, 2, 4, 1, 3)
        dataGrid.addWidget(self.combine_B, 3, 0, 1, 4)
        dataGrid.addWidget(self.split_B, 3, 4, 1, 4)
        
        dataGrid.addWidget(d_divider, 4, 0, 1, -1)
        
        dataGrid.addWidget(reference_label, 5, 0, 1, 6)
        dataGrid.addWidget(self.refSelection, 5, 6, 1, 2)
        
        dataGrid.addWidget(extractPeaksBtn, 6, 0, 1, 4)
        dataGrid.addWidget(self.extractGroupsDialog_Btn, 6, 4, 1, 4)
        dataGrid.addWidget(self.fitHistDialogBtn, 7, 0, 1, 4)
        
        dataPanel.setLayout(dataGrid)
        
        # Baseline controls box
        baseline = QGroupBox("Automatic baseline cleanup")
        base_grid = QGridLayout()
        auto_bs_label = QtGui.QLabel("Baseline removal?")
        auto_bs_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.autobs_Box = pg.ComboBox()
        self.autobs_Box.addItems(['Auto', 'None', 'Lock'])
        self.autobs_Box.setFixedSize(70, 25)
        self.autobs_Box.currentIndexChanged.connect(self.ROI_Change)
        
        # parameters for the auto baseline algorithm
        auto_bs_lam_label = QtGui.QLabel("lambda")
        auto_bs_lam_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.auto_bs_lam_slider = QtGui.QSlider(QtCore.Qt.Horizontal)
        self.auto_bs_lam_slider.setTickPosition(QtGui.QSlider.TicksBothSides)
        self.auto_bs_lam_slider.setMinimum(2)
        self.auto_bs_lam_slider.setMaximum(9)
        self.auto_bs_lam_slider.setValue(6)
        self.auto_bs_lam_slider.setFixedSize(100, 25)
        self.auto_bs_lam_slider.valueChanged.connect(self.ROI_Change)
        
        auto_bs_P_label = QtGui.QLabel("p")
        auto_bs_P_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.auto_bs_P_slider = QtGui.QSlider(QtCore.Qt.Horizontal)
        self.auto_bs_P_slider.setMinimum(0)
        self.auto_bs_P_slider.setMaximum(20)
        self.auto_bs_P_slider.setTickPosition(QtGui.QSlider.TicksBothSides)
        self.auto_bs_P_slider.setValue(3)
        self.auto_bs_P_slider.setFixedSize(100, 25)
        self.auto_bs_P_slider.valueChanged.connect(self.ROI_Change)
        
        # Savitsky-Golay smoothing is very aggressive and doesn't work well in this case
        SGsmoothing_label = QtGui.QLabel("Savitzky-Golay smoothing")
        SGsmoothing_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        SGsmoothing_label.setFixedWidth(160)
        self.SGsmoothing_CB = pg.ComboBox()
        self.SGsmoothing_CB.setFixedSize(70, 25)
        self.SGsmoothing_CB.addItems(['Off','On'])
        self.SGsmoothing_CB.currentIndexChanged.connect(self.ROI_Change)
        
        SG_window_label = QtGui.QLabel("Window")
        SG_window_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        
        self.SGWin_Spin = pg.SpinBox(value=15, step=2, bounds=[5, 49], delay=0, int=True)
        self.SGWin_Spin.setFixedSize(60, 25)
        self.SGWin_Spin.valueChanged.connect(self.ROI_Change)
        
        # should be inactive until extraction
        self.save_baselined_ROIs_Btn = QtGui.QPushButton('Save baselined ROI traces')
        self.save_baselined_ROIs_Btn.clicked.connect(self.save_baselined)
        self.save_baselined_ROIs_Btn.setDisabled(True)
        
        b_divider = QHLine()
        b_divider.setFixedWidth(375)
        
        base_grid.addWidget(auto_bs_label, 0, 0, 1, 2)
        base_grid.addWidget(self.autobs_Box, 0, 2, 1, 1)
        base_grid.addWidget(self.save_baselined_ROIs_Btn, 1, 0, 1, 3)
        base_grid.addWidget(auto_bs_P_label, 0, 3)
        base_grid.addWidget(self.auto_bs_P_slider, 0, 4, 1, 2)
        base_grid.addWidget(auto_bs_lam_label, 1, 3)
        base_grid.addWidget(self.auto_bs_lam_slider, 1, 4, 1, 2)
        
        base_grid.addWidget(b_divider, 2, 0, 1, -1)
        
        base_grid.addWidget(SGsmoothing_label, 3, 0, 1, 3)
        base_grid.addWidget(self.SGsmoothing_CB, 3, 3)
        base_grid.addWidget(SG_window_label, 3, 4)
        base_grid.addWidget(self.SGWin_Spin, 3, 5)
    
        baseline.setLayout(base_grid)
        
        
        # peak finding controls box
        peakFinding = QGroupBox("Peak finding and editing")
        pkF_grid = QGridLayout()
        
        # Switch for manual peak finding
        self.manual = QRadioButton("Edit peaks with mouse", self)
        self.auto = QRadioButton("Auto find peaks", self)
        self.man_auto_group = QButtonGroup()
        self.man_auto_group.addButton(self.manual)
        self.man_auto_group.addButton(self.auto)
        
        if self.autoPeaks:
            self.auto.setChecked(True)
            
        self.man_auto_group.buttonClicked.connect(self.manualPeakToggle)
        
        p_divider = QHLine()
        p_divider.setFixedWidth(375)
        
        p3_show_label = QtGui.QLabel("Show")
        p3_show_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        p3_select_label = QtGui.QLabel("in zoom.")
        self.p3Selection = QtGui.QComboBox()
        self.p3Selection.setFixedSize(90, 25)       # only width seems to work
        self.p3Selection.addItems(['-'])
        self.p3Selection.currentIndexChanged.connect(self.ROI_Change)
                
        # Toggle between wavelet transform and simple algorithm for peak finding
        peakFind_L_label = QtGui.QLabel("Auto-find peaks with")
        peakFind_L_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        peakFind_R_label = QtGui.QLabel("algorithm.")
        peakFind_R_label.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        cwt_width_label = QtGui.QLabel("Width (wavelet only)")
        cwt_width_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        SNR_label = QtGui.QLabel("Prominence / SNR")
        SNR_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        
        self.peak_CB = pg.ComboBox()
        self.peak_CB.setFixedSize(90, 25)
        self.peak_CB.addItems(['wavelet','simple'])
        self.peak_CB.currentIndexChanged.connect(self.ROI_Change)
        
        # spin boxes for CWT algorithm parameters
        self.cwt_SNR_Spin = pg.SpinBox(value=1.3, step=.1, bounds=[.1, 4], delay=0, int=False)
        self.cwt_SNR_Spin.setFixedSize(70, 25)
        self.cwt_SNR_Spin.valueChanged.connect(self.ROI_Change)
        
        self.cwt_w_Spin = pg.SpinBox(value=6, step=1, bounds=[2, 20], delay=0, int=True)
        self.cwt_w_Spin.setFixedSize(70, 25)
        self.cwt_w_Spin.valueChanged.connect(self.ROI_Change)
        
        # Control to exclude small peaks
        removeSml_L_label = QtGui.QLabel("Ignore peaks < ")
        removeSml_L_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        removeSml_R_label = QtGui.QLabel(" of largest.")
        removeSml_R_label.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.removeSml_Spin = pg.SpinBox(value=30, step=10, bounds=[0, 100], suffix='%', delay=0, int=False)
        self.removeSml_Spin.setFixedSize(70, 25)
        self.removeSml_Spin.valueChanged.connect(self.ROI_Change)
        
        # should be inactive until peaks are extracted
        self.savePSRBtn = QtGui.QPushButton('Save peak data')
        self.savePSRBtn.clicked.connect(self.save_peaks)
        self.savePSRBtn.setDisabled(True)
        
        showDataBtn = QtGui.QPushButton('Show current peak data')
        showDataBtn.clicked.connect(self.resultsPopUp)
        
        pkF_grid.addWidget(p3_show_label, 0, 0)
        pkF_grid.addWidget(p3_select_label, 0, 2, 1, 2)
        pkF_grid.addWidget(self.p3Selection, 0 , 1)
        
        pkF_grid.addWidget(p_divider, 1, 0, 1, -1)
        
        pkF_grid.addWidget(self.manual, 2, 2, 1, 2)
        pkF_grid.addWidget(self.auto, 2, 0, 1, 2)
        
        pkF_grid.addWidget(peakFind_L_label, 3, 0, 1, 2)
        pkF_grid.addWidget(self.peak_CB, 3, 2)
        pkF_grid.addWidget(peakFind_R_label, 3, 3)
        pkF_grid.addWidget(SNR_label, 4, 0, 1, 2)
        pkF_grid.addWidget(self.cwt_SNR_Spin, 4, 2)
        pkF_grid.addWidget(cwt_width_label, 5, 0, 1, 2)
        pkF_grid.addWidget(self.cwt_w_Spin, 5, 2)
        pkF_grid.addWidget(removeSml_L_label, 6, 0, 1, 2)
        pkF_grid.addWidget(self.removeSml_Spin, 6, 2)
        pkF_grid.addWidget(removeSml_R_label, 6, 3)
        
        pkF_grid.addWidget(showDataBtn, 7, 0, 1, 2)
        pkF_grid.addWidget(self.savePSRBtn, 7, 2, 1, 2)
        
        pkF_grid.setSpacing(10)
        #pkF_grid.setColumnStretch(0,3)
        #pkF_grid.setColumnStretch(2,2)
        peakFinding.setLayout(pkF_grid)
        
    
        #stack widgets into control panel
        dataPanel.setFixedHeight(225)
        baseline.setFixedHeight(150)
        peakFinding.setFixedHeight(225)
        histograms.setFixedHeight(150)
        
        controls.addWidget(dataPanel, 0, 0, 2, -1)
        controls.addWidget(baseline, 2, 0 , 1, -1)
        controls.addWidget(peakFinding, 3, 0 , 2, -1)
        controls.addWidget(histograms, 5 , 0, 1, -1)
        
        controls.setFixedWidth(420)
        
        self.central_layout.addWidget(controls, 0, 3, -1, 1)
        return
    
    def updateDatasetComboBox(self, _name):
        """Return value indicates a duplicate name was found"""
        #self.evasion = False
        print ("self.datasetListCBX {}".format(self.datasetList_CBX))
        if self.datasetList_CBX == ['-']:
            # the list is empty so reset with the passed value
            self.datasetCBx.setItems([_name])
            self.datasetList_CBX = [_name]
            return False
        else:
            # add new data set to combobox
            if _name in self.datasetList_CBX:
                # get random 3 letter string and add it
                _s = utils.getRandomString(3)
            
                _sname = _name + _s
                self.datasetCBx.addItem(_sname)
                self.datasetList_CBX.append(_sname)
                #self.evasion = True
                return _sname
            else:
                self.datasetCBx.addItem(_name)
                self.datasetList_CBX.append(_name)
              
                return False
       
    
    def resultsPopUp(self):
        """Make a pop up window of the current peak results"""
        _ROI = self.ROI_selectBox.currentText()
        _r = self.workingDataset.resultsDF.df[_ROI]
        #print (_r, type(_r))
        qmb = QDialog()
        qmb.setWindowTitle('Peaks from {}'.format(_ROI))
        qmb.setGeometry(1000,600,600,800)
        self.peaksText = QtGui.QTextEdit()
        font = QtGui.QFont()
        font.setFamily('Courier')
        font.setFixedPitch(True)
        font.setPointSize(12)
        self.peaksText.setCurrentFont(font)
        self.peaksText.setText(_r.to_string())
        self.peaksText.setReadOnly(True)
        
        #add buttons, make it the right size
        qmb.layout = QVBoxLayout()
        qmb.layout.addWidget(self.peaksText)
        qmb.setLayout(qmb.layout)
        qmb.exec_()
    
    def histogram_parameters(self):
        _nbins = int(self.histo_NBin_Spin.value())
        _max = self.histo_Max_Spin.value()
        return _nbins, _max
    
       
    def doHistograms(self):
        """called for histogram output"""
        _nbins, _max = self.histogram_parameters()
        _condList = self.conditions + ["Sum"]
        
        # create a dataframe to put the results in
        self.hDF = HistogramsR(self.workingDataset.ROI_list, _condList, _nbins, 0., _max)
        
        maxVal = len (self.workingDataset.ROI_list) * len (_condList)
        progMsg = "Histograms for {0} traces".format(maxVal)
        
        with pg.ProgressDialog(progMsg, 0, maxVal) as dlg:
            
            #from the allowlist, should be from edited internal data?
            for _condi, _pdf in self.gpd.pk_extracted_by_condi.items():
                
                print (_condi, _pdf.columns)
                for _ROI in _pdf.columns:
                    dlg += 1
                    # calculate individual histograms and add to dataframe
                    hy, hx = np.histogram(_pdf[_ROI], bins=_nbins, range=(0., _max))
                    self.hDF.addHist(_ROI, _condi, hy)
                    
        # add sum columns
        self.hDF.ROI_sum()
    
    def updateHistograms(self):
        """called when histogram controls are changed"""
        
        # get controls values and summarise to terminal
        _nbins, _max = self.histogram_parameters()
        _ROI = self.ROI_selectBox.currentText()
        _hsum = self.sum_hist.currentText()
        print ('Update {3} Histogram(s) for {2} with Nbins = {0} and maximum dF/F = {1}.'.format(_nbins, _max, _ROI, _hsum))
       
        # clear
        self.p2.clear()
        
        if _hsum == "Separated":
            for i, _condi in enumerate(self.conditions):
                # colours
                col_series = (i, len(self.conditions))
                # get relevant peaks data for displayed histograms
                _, _pdata = self.workingDataset.resultsDF.getPeaks(_ROI, _condi)
                # redo histogram
                hy, hx  = np.histogram(_pdata, bins=_nbins, range=(0., _max))
                # replot
                self.p2.plot(hx, hy, name="Histogram "+_condi, stepMode=True, fillLevel=0, pen=col_series, brush=col_series) ###fillOutline=True,
        
        elif _hsum == "Summed":
            sumhy = np.zeros(_nbins)
            for _condi in self.conditions:
                _, _pdata = self.workingDataset.resultsDF.getPeaks(_ROI, _condi)
                hy, hx  = np.histogram(_pdata, bins=_nbins, range=(0., _max))
                sumhy += hy
            
            self.p2.plot(hx, sumhy, name="Summed histogram "+_ROI, stepMode=True, fillLevel=0, fillOutline=True, brush='y')
            
            if self.fitHistogramsOption:
                #print ("len hx {}, hy {}".format(len(hx), len(hy)))
                _num = self.histo_nG_Spin.value()
                _q = self.histo_q_Spin.value()
                _ws = self.histo_Max_Spin.value() / 20
                
                _hxc = np.mean(np.vstack([hx[0:-1], hx[1:]]), axis=0)
                _opti = fit_nGaussians(_num, _q, _ws, sumhy, _hxc)
                if _opti.success:
                    _hx_u, _hy_u = nGaussians_display (_hxc, _num, _opti.x)
                    _qfit = _opti.x[0]
                    _c = self.p2.plot(_hx_u, _hy_u, name='Fit of {} Gaussians q: {:.2f}'.format(_num,_qfit))
                    #from pyqtgraph.examples
                    _c.setPen('w', width=3)
                    _c.setShadowPen(pg.mkPen((70,70,30), width=8, cosmetic=True))
                else:
                    print ("fit failed")
        
       
    def datasetChange(self):
        print ("a (dataset) change is coming")
        
        if self.datasetCBx.currentText() != self.workingDataset.DSname:
            # prep current data for store
            # store GUI settings?
            self.store.storeSet(copy.copy(self.workingDataset))
            print ('Stored {}'.format(self.workingDataset.DSname))
            
            self.workingDataset = self.store.retrieveWorkingSet(self.datasetCBx.currentText())
            print ('Retrieved {}'.format(self.workingDataset.DSname))
        
            # print GUI control dict
            print ("swdsGC {}".format(self.workingDataset.GUIcontrols))
            
            # do this first otherwise on ROI extracted peaks are overwritten
            # execute GUI controls specified in the retrieved Dataset
            for k,v in self.workingDataset.GUIcontrols.items():
                if k == "autoPeaks":
                    self.autoPeaks_GUI_switch(v)
                elif k == "print":
                    print (v)
            
            # update the ROI list combobox (should have a list)
            # needed for updating GUI
            self.updateROI_list_Box()
            self.ROI_Change()
            
            
        
  
    def autoPeaks_GUI_switch(self, v):
        if v == "Disable":
            print ("autoPeaks_GUI_switch : Disable")
            self.pauseUpdates = True
            self.autobs_Box.setValue('None')
            self.auto_bs = False
            self.manual.setChecked(True)
            self.manual.setDisabled(True)
            self.auto.setDisabled(True)
            self.autoPeaks = False
            
            self.pauseUpdates = False
            self.manualPeakToggle()
            
        elif v == "Enable":
            print ("autoPeaks_GUI_switch : Enable")
            #self.autobs_Box.setValue('Auto')
            #self.auto_bs = True
            #self.manual.setChecked(False)
            self.pauseUpdates = True
            self.manual.setEnabled(True)
            self.auto.setEnabled(True)
            #self.autoPeaks = True
            self.pauseUpdates = False
            self.manualPeakToggle()
    
    def getGroups(self):
        """launch group processing dialog"""
        print ('Process grouped peaks from all ROIs.')
        self.getgroupsDialog = groupPeakDialog()
        _dataset = copy.copy(self.workingDataset)
        ddf = utils.decomposeRDF(_dataset.resultsDF.df)
        self.getgroupsDialog.addData(ddf, name=_dataset.DSname)
        accepted = self.getgroupsDialog.exec_()
      
    def launchHistogramFit(self):
        """Wrapping function to launch histogram fit dialog"""
        print ('Dialog to obtain quantal parameters from histogram fits.')
        self.hfd = histogramFitDialog()
        #send current peak data
        _dataset = copy.copy(self.workingDataset)
        
        ddf = utils.decomposeRDF(_dataset.resultsDF.df)
        self.hfd.addData(ddf, _dataset.DSname, _dataset.getSD(maskWidth=10))
        
        accepted = self.hfd.exec_()
        
    def extractAllPeaks(self):
        """Wrapping function to get peak data from the dialog"""
        print ('Dialog for getting peaks from all ROIs according to reference pattern.')
        # if the QDialog object is instantiated in __init__, it persists in state....
        # do it here to get a fresh one each time.
        self.gpd = extractPeaksDialog()
        
        # pass the data into the get peaks dialog object
        # we do not want the original trace data modified
        _dataset = copy.copy(self.workingDataset)
        
        # automatically reduce baseline (could also do this interactively??)
        # baselineIterator includes a progress indicator.
        
        # can be Auto or Lock (meaning GUI controls are not updating algorithm)
        if self.autobs_Box.value() != 'None':
            
            if self.autobs_Box.value() == 'Auto':
                # populate values for automatic baseline removal from GUI (unless 'Lock')
                self.setBaselineParams()
            
            _dataset.traces = baselineIterator(_dataset.traces, self.auto_bs_lam, self.auto_bs_P)
        
        #get the times of the peaks from the "best" trace, that were selected auto or manually
        _peak_t, _ = self.workingDataset.resultsDF.getPeaks('Mean', self.refSelection.currentText())  # pd.series
        
        #print ("srsv: {}".format(self.refSelection.currentText()))
        
        _sorted_peak_t = _peak_t.sort_values(ascending=True)    # list is not sorted until now
        _sorted_peak_t.dropna(inplace=True)                     # if there are 'empty' NaN, remove them
        
        self.workingDataset.resultsDF.peakTimes = _sorted_peak_t # the definitive list of peak times, woud be degraded by editing but will only be used for masking traces later.
        
        print ("swdrdfpt : {}".format(self.workingDataset.resultsDF.peakTimes))
        
        self.extPa["tPeaks"] = _sorted_peak_t
        
        # pass in "external parameters" for the peak extraction via extPa
        self.gpd.setExternalParameters(self.extPa)
        
        #reordered because peaks must be put first. 
        self.gpd.addDataset(_dataset)
        
        # returns 1 (works like True) when accept() or 0 (we take for False) otherwise.
        # data from dialog is stored in the attributes of self.gpd
        accepted = self.gpd.exec_()
        
        if accepted:
            self.noPeaks = False

            print (self.gpd.pk_extracted_by_condi) #will be the allow list if the allowlist
            # these should now become available to be viewed (even edited?)
            
            #make 'save' and other analysis buttons available
            self.fitHistDialogBtn.setEnabled(True)
            self.savePSRBtn.setEnabled(True)
            self.save_baselined_ROIs_Btn.setEnabled(True)
            self.extractGroupsDialog_Btn.setEnabled(True)
        
            # create new data set
            extracted = Dataset(self.gpd.name)
          
            # specify GUI state for extracted peaks
            #extracted.GUIcontrols['print'] = 'here is a GUI command'
            extracted.GUIcontrols["autoPeaks"] = 'Disable'
            
            # update combobox
            _duplicate = self.updateDatasetComboBox(str(extracted.DSname))
            if _duplicate:
                extracted.setDSname(_duplicate)                 #fix name if it was a duplicate
                print ("duplicate name {}".format(_duplicate))  # add results to new set
            
            # add the extracted peaks to a resultsDF instance and place that in the dataset
            _resdf = self.gpd.pk_extracted_by_condi
            extracted_peaksRDF = Results()                      # generate empty resultsDF object
            extracted_peaksRDF.addPeaksExtracted(_resdf)        # conversion
            extracted.addPeaksToDS (extracted_peaksRDF)
            extracted.ROI_list = copy.copy(extracted_peaksRDF.ROI_list)
            
            _bl_resd = self.gpd.excludedListedByCondi
            print ("bl_resd {}".format(_bl_resd))
            
            if self.gpd.excludedCount == 0:
                extracted.excludedList = None
                print ("No traces on excluded list.")
            else:
                len(_bl_resd)
                print ("{} excluded traces for bad SNR".format(self.gpd.excludedCount))
                _bl_prdf = Results()
                _bl_prdf.addPeaksExtracted(_bl_resd)
                extracted.excludelisted = _bl_prdf
            
            # add baselined traces to new set
            extracted.addTracesToDS(self.gpd.tracedata)
            extracted.peakTimes = _sorted_peak_t            # the same as the source dataset (for masking)
            
            # store
            self.store.storeSet(extracted)
            
            #for convenience, switch immediately to display the dataset that was just created
            self.datasetCBx.setValue(extracted.DSname)
            
        
        else:
            print ('Returned but not happily: self.gpd.pk_extracted_by_condi is {}'.format(self.gpd.pk_extracted_by_condi))
            
            # displaying output would make no sense
            
        #ideas:
        
        # accumulate histogram from individual ROI or store separately
        
    def plotNewData(self):
        """Do some setup immediately after data is loaded"""
        
        _sel_condi = self.p3Selection.currentText()
        print ("Plot New Data with the p3 selector set for: ", _sel_condi)
        y = {}
        
        self.p1.clear()
        self.p3.clear()
        
        for i, _condi in enumerate(self.conditions):
            x = self.workingDataset.traces[_condi].index
            y[i] = self.workingDataset.traces[_condi].mean(axis=1).to_numpy()

            self.p1.plot(x, y[i], pen=(i,3))
        
            if _sel_condi == _condi:
                # curve
                self.p3.plot(x, y[i], pen=(i,3))
                
                if self.autoPeaks:
                    xp, yp = self.peaksWrapper(x, y[i], _condi)
                    
                else:
                    # if new data is loaded without autopeaks, there are no peaks....
                    # this goes wrong later on though
                    xp = np.array([])
                    yp = np.array([])
                
                # need to add something to p3 scatter
                self.p3.plot(xp, yp, name="Peaks "+_condi, pen=None, symbol="s", symbolBrush=(i,3))
                self.plots.peakslabel.setText("{} peaks in {} condition".format(len(yp), _condi))
                
                # create the object for parsing clicks in p3
                self.cA = clickAlgebra(self.p3)
                _p3_scatter = utils.findScatter(self.p3.items)
                if _p3_scatter:
                    _p3_scatter.sigClicked.connect(self.clickRelay)
                    _p3_scatter.sigPlotChanged.connect(self.manualUpdate)
        
        self.createLinearRegion()
        #return
        
    def findSimplePeaks(self, xdat, ydat, name='unnamed'):
        """Simple and dumb peak finding algorithm"""
        # cut_off is not implemented here
        # SNR is used as a proxy for 'prominence' in the simple algorithm.
        self.cwt_SNR = self.cwt_SNR_Spin.value()
        
        peaks, _ = scsig.find_peaks(ydat, prominence=self.cwt_SNR)
        _npeaks = len(peaks)
        if _npeaks != 0:
            print ('Simple peak finding algorithm found {0} peaks in {1} trace with prominence {2}'.format(_npeaks, name, self.cwt_SNR))
            
            xp = xdat[peakcwt]
            yp = ydat[peakcwt]
            
        else:
            print ('No peaks found in {0} trace with simple algorithm with prominence {1}'.format(name, self.cwt_SNR))
            
            xp = []
            yp = []
           
        return xp, yp
        
        
    def findcwtPeaks(self, xdat, ydat, name='unnamed'):
        """Find peaks using continuous wavelet transform"""
        # indices in peakcwt are not zero-biased
        self.cwt_width = self.cwt_w_Spin.value()
        self.cwt_SNR = self.cwt_SNR_Spin.value()
        peakcwt = scsig.find_peaks_cwt(ydat, np.arange(1, self.cwt_width), min_snr=self.cwt_SNR) - 1
        _npeaks = len(peakcwt)
        if _npeaks != 0:
            xpeak = xdat[peakcwt]
            ypeak = ydat[peakcwt]
            
            # filter out small peaks
            _cutOff = float (self.removeSml_Spin.value()) * ydat.max() / 100.0
            xpf = xpeak[np.where(ypeak > _cutOff)]
            ypf = ypeak[np.where(ypeak > _cutOff)]
            
            print ('wavelet transform peak finding algorithm found {0} peaks in {1} trace, width: {2}, SNR: {3}, cutOff: {4}.'.format(_npeaks, name, self.cwt_width, self.cwt_SNR, _cutOff))
        else:
            print ('No peaks found in {0} with cwt algorithm, width: {1}, SNR: {2}, cutOff: {4}.'.format(name, self.cwt_width, self.cwt_SNR, _cutOff))
            xpf = []
            ypf = []
        #_condi = self.p3Selection.currentText() == _condi:
        #self.plots.peakslabel.setText("{} peaks in ".format(_npeaks, _condi))
        
        return xpf, ypf
    
    def manualUpdate(self):
        """Some editing was done in p3, so update other windows accordingly"""
        print ('Peak data in p3 changed manually')
       
        _sel_condi = self.p3Selection.currentText()
        _ROI = self.ROI_selectBox.currentText()
        
        # update the peaks in p1 and histograms only
        utils.removeAllScatter(self.p1, verbose=False)
        
        #update p2 histograms
        self.updateHistograms()
        
        for i, _condi in enumerate(self.conditions):
            #colours
            col_series = (i, len(self.conditions))
            
            if _sel_condi == _condi :
                _scatter = utils.findScatter(self.p3.items)
                # sometimes a new scatter is made and this "deletes" the old one
                # retrieve the current manually curated peak data
                if _scatter is None:
                    print ('No Scatter found, empty data.')
                    xp = []
                    yp = []
                else:
                    xp, yp = _scatter.getData()
                
                # write peaks into results
                self.workingDataset.resultsDF.addPeaks(_ROI, _sel_condi, xp, yp)
                # print (self.workingDataset.resultsDF.df[_ROI])
             
            xp, yp = self.workingDataset.resultsDF.getPeaks(_ROI, _condi)
            
            if self.split_traces:
                _target = self.p1stackMembers[i]
                # only one scatter item in each split view
                _t_scat = utils.findScatter(_target.items)
                _t_scat.setData(xp, yp, brush=col_series)
                
            else:
                self.p1.plot(xp, yp, pen=None, symbol="s", symbolBrush=col_series)
            
            self.plots.peakslabel.setText("{} peaks in {} condition.".format(len(yp), _condi))
                
    def setBaselineParams (self):
        """Get parameters for auto baseline from GUI"""
        
        self.auto_bs_lam =  10 ** self.auto_bs_lam_slider.value()
        self.auto_bs_P =  10 ** (- self.auto_bs_P_slider.value() / 5)
    
    def peaksWrapper (self, x , y, set):
        """Simplify peak finding calls"""
        
        if self.simplePeaks:
            xp, yp = self.findSimplePeaks(x , y, name=set)
        else:
            xp, yp = self.findcwtPeaks(x , y, name=set)
    
        return xp, yp
        
    def updateROI_list_Box(self):
        """populate the combobox for choosing which ROI to show"""
        self.ROI_selectBox.clear()
        self.ROI_selectBox.addItems(self.workingDataset.ROI_list)
    
    def ROI_Change(self):
        """General 'Update' method"""
        # called when ROI/trace is changed but
        # also when a new peak fit
        # approach is chosen.
        # consider renaming
        
        #### if baseline was changed in another window, all the peaks are now off...
        
        # we are not interested in updating data if there isn't any
        if self.dataLoaded == False:
            return
        # used when the GUI is updating at multiple places
        if self.pauseUpdates:
            return
        # something changed in the control panel, get latest values
        _ROI = self.ROI_selectBox.currentText()
              
        if self.peak_CB.value() == 'simple':
            self.simplePeaks = True
        else:
            self.simplePeaks = False

        if self.autobs_Box.value() != 'None':
            self.auto_bs = True
            # populate values for automatic baseline removal from GUI
            if self.autobs_Box.value() == 'Auto':
                self.setBaselineParams()
        else:
            self.auto_bs = False
            
        if self.SGsmoothing_CB.value() == 'On':
            # populate values for Savitsky-Golay smoothing from GUI
            self.sgSmooth = True
            self.sgWin = self.SGWin_Spin.value()
        else:
            self.sgSmooth = False

        # Empty the trace dictionary and the plots - perhaps we could be more gentle here?
        y = {}
        z = {}
        
        # Rather than doing this, need to keep the peak objects and set their data anew?
        self.p1.clear()
        
        # Rather than clearing objects in p3, we set their data anew
        _p3_items = self.p3.items
        _p3_scatter = utils.findScatter(_p3_items)
        _p3_curve = utils.findCurve(_p3_items)
        
        for i, _condi in enumerate(self.conditions):
            col_series = (i, len(self.conditions))
            x = np.array(self.workingDataset.traces[_condi].index)
            
            if _ROI == "Mean":
                y[i] = self.workingDataset.traces[_condi].mean(axis=1).to_numpy()
                
            elif _ROI == "Variance":
                y[i] = self.workingDataset.traces[_condi].var(axis=1).to_numpy()
                # we never want to subtract the steady state variance
                self.auto_bs = False
                print ('No baseline subtraction for variance trace')
                
            elif _ROI != '':
                print ("condi, roi {} {}".format(_condi, _ROI))
                y[i] = self.workingDataset.traces[_condi][_ROI].to_numpy()
            
            else:
                return
            
            if self.auto_bs:
                # baseline
                z[i] = baseline_als(y[i], lam=self.auto_bs_lam, p=self.auto_bs_P, niter=10)
                
                # subtract the baseline
                y[i] = y[i] - z[i]
                
                # plotting is done below
                
            if self.sgSmooth:
                print ('Savitsky Golay smoothing with window: {0}'.format(self.sgWin))
                y[i] = savitzky_golay(y[i], window_size=self.sgWin, order=4)

            if self.autoPeaks:
                
                # call the relevant peak finding algorithm
                xp, yp = self.peaksWrapper(x, y[i], _condi)
                self.plots.peakslabel.setText("{} peaks in {} condition".format(len(yp), _condi))
                
                # write automatically found peaks into results
                self.workingDataset.resultsDF.addPeaks(_ROI, _condi, xp, yp)
                
                
            else: # we are in manual peaks
                
                # read back existing peak data from results (might be empty if it's new ROI)
                xp, yp = self.workingDataset.resultsDF.getPeaks(_ROI, _condi)
                if len(yp) == 0: print ("Peak results for {} {} are empty".format( _ROI, _condi))
                else :
                    try:
                        print ("Retrieved: {} {} first xp,yp : {}, {}".format( _ROI, _condi, xp[0], yp[0]))
                    except KeyError:
                        #print (xp.shape, yp.shape)
                        print ("No peaks, xp or yp empty? {} {} {} {}".format(_ROI, _condi, xp, yp))
                        
            # draw p1 traces and scatter
            if self.split_traces:
                target = self.p1stackMembers[i]
                target.clear()
                target.plot(x, y[i], pen=col_series)
                if len(yp) > 0 : target.plot(xp, yp, pen=None, symbol="s", symbolBrush=col_series)
            else:
                self.p1.plot(x, y[i], pen=col_series)
                if len(yp) > 0 : self.p1.plot(xp, yp, pen=None, symbol="s", symbolBrush=col_series)
                
                #plot baseline, offset by the signal max.
                if self.auto_bs:
                    self.p1.plot(x, z[i]-y[i].max(), pen=(255,255,255,80))
            
            #p3: plot only the chosen trace
            if self.p3Selection.currentText() == _condi:
                self.plots.peakslabel.setText("{} peaks in {} condition.".format(len(yp), _condi))
                if _p3_scatter is None:
                    # Do something about it, there is no peak scatter yet in this graph
                    if len(yp) > 0 :
                        self.p3.plot(xp, yp, name="Peaks "+_condi, pen=None, symbol="s", symbolBrush=col_series)
                else:
                    _p3_scatter.clear()
                    if len(yp) > 0 : _p3_scatter.setData(xp, yp, brush=col_series)
                
                _p3_curve.clear()
                _p3_curve.setData(x, y[i], pen=col_series)
                
        self.createLinearRegion()
        
        self.updateHistograms()
        
        return
        
    def setRanges(self):
        """ Collect the extremities of data over a set of conditions """
        self.ranges = {}
        # use the first condition (sheet) as a basis
        _df = self.workingDataset.traces[self.conditions[0]]
        self.ranges['xmin'] = _df.index.min()
        self.ranges['xmax'] = _df.index.max()
        self.ranges['ymin'] = _df.min().min()
        self.ranges['ymax'] = _df.max().max()
        
        # lazily compare across all conditions (including the first)
        for sheet in self.workingDataset.traces.values():
            if sheet.min().min() < self.ranges['ymin']:
                self.ranges['ymin'] = sheet.min().min()
            if sheet.max().max() > self.ranges['ymax']:
                self.ranges['ymax'] = sheet.max().max()
                
            if sheet.index.min() < self.ranges['xmin']:
                self.ranges['xmin'] = sheet.index.min()
            if sheet.index.max() > self.ranges['xmax']:
                self.ranges['xmax'] = sheet.index.max()
        return
    
    def save_peaks(self):
        print ("save extracted peak data and optionally histograms")
        
        # needs to be updated for openpyxl
        #format for header cells.
        #self.hform = {
        #'text_wrap': True,
        #'valign': 'top',
        #'fg_color': '#D5D4AC',
        #'border': 1}
        
        if self.noPeaks:        #nothing to save
            print ('Nothing to save, no peaks found yet!')
            return
        
        self.filename = QFileDialog.getSaveFileName(self,
        "Save Peak Data", os.path.expanduser("~"))[0]
        
        if self.filename:
            _wds = self.workingDataset
            wb = Workbook()
            
            # combine allowlist and excludelist dictionaries for output
            #_output = {**self.gpd.pk_extracted_by_condi, **self.gpd.excludelisted_by_condi}
            _allowedPeaks = {}
            _excludedPeaks = {}
            
            if _wds.resultsDF:
                _allowedPeaks = utils.decomposeRDF (_wds.resultsDF.df)
            if _wds.excludelisted:
                _excludedPeaks = utils.decomposeRDF (_wds.excludelisted.df)
            
            _output = {**_allowedPeaks, **_excludedPeaks}
            
            for _condi, _resultdf in _output.items():
                # in case there are duplicate peaks extracted, remove them and package into dummy variable
                # this syntax means : loc["not" the duplicates]
                _pe = _resultdf.loc[~_resultdf.index.duplicated(keep='first')] #StackOverflow 13035764
                
                wb.create_sheet(_condi)
                _wcs = wb[_condi]
                
                # write customised header
                for _num, _pcol in enumerate(_pe.columns.values):
                    _wcs.cell(1, _num + 1).value = _pcol + " " + _condi      #, header_format)
                    
                # write out data
                for _row in dataframe_to_rows(_pe, index=False, header=False):
                    _wcs.append(_row)
            
                #write index
                _wcs.insert_cols(1)
                for _num, _pin in enumerate(_pe.index.values):
                    _wcs.cell(1,1).value = "Time"
                    _wcs.cell(_num + 2, 1).value = _pin
            
            wb.remove(wb['Sheet'])
            
            if self.saveHistogramsOption:
                wb = self.save_histograms(wb)
            
            wb.save(self.filename)
            print ("Saved peaks from to workboook {}".format(self.filename))
          
            

    def save_histograms(self, wb):
        """write out histograms for each ROI to excel workbook"""
        
        print ("saving histograms")
        self.doHistograms()
        print (self.hDF.df.head(5))
        #save histograms into new sheet
        wb.create_sheet("Histograms")
        _wcs = wb["Histograms"]
    
        # write header
        for col_num, col in enumerate(self.hDF.df.columns.values):
            _wcs.cell(1, col_num + 1).value = str(col) + " hi"
        
        # histogrammed data
        for _row in dataframe_to_rows(self.hDF.df, index=False, header=False):
            _wcs.append(_row)
        
        # bin edges go in first column
        _wcs.insert_cols(1)
        for _num, _pin in enumerate(self.hDF.binEdges):
            _wcs.cell(1,1).value = "BinEdges"
            _wcs.cell(_num + 2, 1).value = _pin
            
        return wb
        
    def save_baselined(self):
        # save baselined traces
        # No filtering so far
        print ("save_baselined data?")
        self.btfilename = QFileDialog.getSaveFileName(self,
        "Save Baselined ROI Data", os.path.expanduser("~"))[0]
        
        if self.btfilename:
            _wdsT = self.workingDataset.traces
            wb = Workbook()
            for _condi, _tdf in _wdsT.items():
                wb.create_sheet(_condi)
                _wcs = wb[_condi]
                
                # write customised header
                for _num, _pcol in enumerate(_tdf.columns.values):
                    _wcs.cell(1, _num + 1).value = _pcol + " " + _condi      #, header_format)
                    
                # write out data
                for _row in dataframe_to_rows(_tdf, index=False, header=False):
                    _wcs.append(_row)
            
                #write index
                _wcs.insert_cols(1)
                for _num, _pin in enumerate(_tdf.index.values):
                    _wcs.cell(1,1).value = "Time"
                    _wcs.cell(_num + 2, 1).value = _pin
            
            # was created at start
            
            wb.remove(wb['Sheet'])
            wb.save(self.btfilename)
            print ("Saved traces from to workboook {}".format(self.btfilename))
            
    def open_file(self):
        """Open a dialog to provide sheet names"""
        
        self.filename = QFileDialog.getOpenFileName(self,
            "Open Data", os.path.expanduser("~"))[0]
        
        if self.filename:
            """   #OLD WAY
                #very simple and rigid right now - must be an excel file with conditions
                #should be made generic - load all conditions into dictionary of dataframes no matter what
            with pg.ProgressDialog("Loading conditions...", 0, len(self.conditions)) as dlg:
                _traces = {}
                for _sheet in self.conditions:
                    dlg += 1
                    try:
                        _traces[_sheet] = pd.read_excel(self.filename, sheet_name=_sheet, index_col=0)
                        print ("XLDR: From spreadsheet- {}\n{}".format(_sheet, _traces[_sheet].head()))
                    except:
                        print ("Probably: XLDR error- no sheet named exactly {0}. Please check it.".format(_sheet))
                        self.conditions.remove(_sheet)
                # decide if there is data or not
            """
            
            # Now as a oneliner
            #"None" reads all the sheets into a dictionary of data frames
            _traces = pd.read_excel(self.filename, None, index_col=0)
        
        else:
            print ("file dialog failed")
            return
        
        self.conditions = list(_traces.keys())
        
        print ("Loaded following conditions: ", self.conditions)
        
        if self.workingDataset.isEmpty:
            print ("First data set loaded")
        
        else:
            #store existing working dataset
            self.store.storeSet(copy.copy(self.workingDataset))
            print ("Putting {} in the store.".format(self.workingDataset.DSname))
        
        # overwrite current working set
        self.workingDataset.addTracesToDS(_traces)
        self.workingDataset.isEmpty = False
        _stem = utils.getFileStem(self.filename)
        self.workingDataset.setDSname(_stem)
    
        _DSname = str(self.workingDataset.getDSname())
   
        _duplicate = self.updateDatasetComboBox(_DSname)
        #returns either false or the name to avoid duplicates
        
        #print ("4 {}".format(self.workingDataset.__dict__))
        if _duplicate:
            # update
            self.workingDataset.DSname = _duplicate
        
        self.workingDataset.ROI_list = ["Mean", "Variance"]
        
        _first = self.conditions[0]
        # print (self.workingDataset.__dict__)
        
        # by default use the last sheet, but allow the user to change it
        self.conditionForExtraction = self.conditions[-1]
        print ("scfe: {}".format(self.conditionForExtraction))
        
        self.workingDataset.ROI_list.extend(self.workingDataset.traces[_first].columns.tolist())
        self.updateROI_list_Box()
        
        #find out and store the size of the data
        self.setRanges()
        
        #split trace layout can be made now we know how many sets (conditions) we have
        self.createSplitTraceLayout()
        
        # populate the comboboxes for choosing the data shown in the zoom view,
        # and choosing the reference ROI for peak extraction
        self.p3Selection.clear()
        self.p3Selection.addItems(self.conditions)
        self.refSelection.clear()
        self.refSelection.addItems(self.conditions)
        
        i = self.refSelection.findText(self.conditionForExtraction)
        if i != -1:
            self.refSelection.setCurrentIndex(i)      # the default
            self.p3Selection.setCurrentIndex(i)     #set both for now
        
        #create a dataframe for peak measurements
        self.workingDataset.resultsDF = Results(self.workingDataset.ROI_list, self.conditions)
        print ("peakResults object created", self.workingDataset.resultsDF, self.workingDataset.ROI_list)
        
        self.plotNewData()
        
        #updates based on GUI can now happen painlessly
        self.dataLoaded = True
        self.ROI_Change()
        


if __name__ == "__main__":
    # Change menubar name from 'python' to 'SAFT' on macOS
    # from https://stackoverflow.com/questions/5047734/
    if sys.platform.startswith('darwin'):
    # Python 3: pyobjc-framework-Cocoa is needed
        try:
            from Foundation import NSBundle
            bundle = NSBundle.mainBundle()
            if bundle:
                app_name = os.path.splitext(os.path.basename(sys.argv[0]))[0]
                app_info = bundle.localizedInfoDictionary() or bundle.infoDictionary()
                if app_info:
                    app_info['CFBundleName'] = app_name.upper() # ensure text is in upper case.
        except ImportError:
            print ("Failed to import NSBundle, couldn't change menubar name." )
            
    
    __version__ = "v. 0.4"
    #print (sys.version)
    app = QApplication([])
    smw = SAFTMainWindow()
    smw.show()
    sys.exit(app.exec_())

